import time
import openpyxl
from selenium import webdriver

workbook1 = openpyxl.load_workbook('credentials.xlsx')
sheet = workbook1.active
username = sheet.cell(row=10,column=1).value
password = sheet.cell(row=10,column=2).value

print(f"username:{username},password:{password}")

try:
    driver = webdriver.Chrome()

    driver.get("https://omayo.blogspot.com/")
    driver.delete_all_cookies()
    driver.maximize_window()

    driver.find_element("xpath","//input[@name='userid']").send_keys(str(username))
    driver.find_element("xpath","//input[@name='pswrd']").send_keys(str(password))
    driver.find_element("xpath","//input[@type='submit' or @value='Login']").click()

    time.sleep(5)
except Exception as e:
    print("An error occurred:", e)