import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

print("--- Easy method to read Excel data using openpyxl ---")

# 1. Load the Excel workbook and sheet using openpyxl
workbook = openpyxl.load_workbook('Credentials.xlsx')
sheet = workbook.active  # Selects the first active sheet 

# Get the name of the second sheet (Index 1)
# second_sheet_name = workbook.sheetnames[1]

# Select the sheet directly by its name
# sheet = workbook["Sheet# print(f"Active sheet is now : {sheet.title}")

# Assuming Row 1 has headers (Username, Password)
# Row 2, Column 1 (A2) = Username
# Row 2, Column 2 (B2) = Password
excel_username = sheet.cell(row=2, column=1).value
excel_password = sheet.cell(row=2, column=2).value

print(f"Loaded from Excel -> Username: {excel_username}, Password: {excel_password}")

# 2. Initialize the Chrome Browser
driver = webdriver.Chrome()

try:
    # 3. Open omayo.blogspot.com
    driver.get("https://omayo.blogspot.com/")
    driver.maximize_window()
    time.sleep(2)  # Give the page a moment to load

    # 4. Find the Username and Password input fields
    username_field = driver.find_element(By.NAME, "userid")
    password_field = driver.find_element(By.NAME, "pswrd")

    # 5. Type the credentials retrieved via openpyxl
    username_field.send_keys(str(excel_username))
    password_field.send_keys(str(excel_password))

    # 6. Find the login button and click it
    login_button = driver.find_element(By.XPATH, "//input[@type='submit' or @value='Login']")
    login_button.click()

    print("Form submitted successfully!")

except Exception as e:
    print("An error occurred:", e)

# print("--- loop through data, Starting Data-Driven Testing with openpyxl ---")

# # 1. Load the Excel workbook and sheet using openpyxl
# workbook = openpyxl.load_workbook('credentials.xlsx')
# sheet = workbook.active  # Selects the first active sheet
#
# # 2. Initialize the Chrome Browser
# driver = webdriver.Chrome()
#
# try:
#     # 3. Open omayo.blogspot.com once
#     driver.get("https://omayo.blogspot.com/")
#     driver.maximize_window()
#     time.sleep(2)  # Give the page a moment to load
#
#     # 4. HERE IS WHERE YOU ADD THE LOOP:
#     # We start from row 2 to skip the column headers (Username, Password)
#     # sheet.max_row + 1 ensures the loop includes the very last row
#     for row_idx in range(2, sheet.max_row + 1):
#         excel_username = sheet.cell(row=row_idx, column=1).value
#         excel_password = sheet.cell(row=row_idx, column=2).value
#
#         # Skip empty rows if there are any blank lines at the bottom of the sheet
#         if excel_username is None or excel_password is None:
#             continue
#
#         print(f"Processing Excel Row {row_idx}: Submitting {excel_username}")
#
#         # Find the Username and Password input fields
#         username_field = driver.find_element(By.NAME, "userid")
#         password_field = driver.find_element(By.NAME, "pswrd")
#
#         # Clear any previous text in the boxes before typing new data
#         username_field.clear()
#         password_field.clear()
#
#         # Type the credentials retrieved via openpyxl
#         username_field.send_keys(str(excel_username))
#         password_field.send_keys(str(excel_password))
#
#         # Find the login button and click it
#         login_button = driver.find_element(By.XPATH, "//input[@type='submit' or @value='Login']")
#         login_button.click()
#
#         # Pause for 3 seconds to watch it process before moving to the next row
#         time.sleep(3)
#
#     print("All rows from Excel have been processed successfully using openpyxl!")
#
# except Exception as e:
#     print("An error occurred during execution:", e)
